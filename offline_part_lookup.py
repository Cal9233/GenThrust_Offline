#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import struct
import os
import sys
from datetime import datetime
import json

# Try to import openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed. Excel files will not be read.")

class DBFReader:
    def __init__(self, filename):
        self.filename = filename
        self.records = []
        self.fields = []
        self.read_dbf()
    
    def read_dbf(self):
        try:
            with open(self.filename, 'rb') as f:
                # Read header
                header = f.read(32)
                if len(header) < 32:
                    return
                
                num_records = struct.unpack('<I', header[4:8])[0]
                header_len = struct.unpack('<H', header[8:10])[0]
                record_len = struct.unpack('<H', header[10:12])[0]
                
                # Read field descriptors
                field_data = f.read(header_len - 32)
                field_count = (header_len - 32 - 1) // 32
                
                for i in range(field_count):
                    field_info = field_data[i*32:(i+1)*32]
                    if field_info[0] == 0x0D:  # Field terminator
                        break
                    
                    field_name = field_info[0:11].replace(b'\x00', b'').decode('ascii', errors='ignore').strip()
                    field_type = chr(field_info[11])
                    field_length = field_info[16]
                    
                    self.fields.append({
                        'name': field_name,
                        'type': field_type,
                        'length': field_length
                    })
                
                # Skip to records
                f.seek(header_len)
                
                # Read records
                for _ in range(min(num_records, 50000)):  # Limit records for performance
                    record_data = f.read(record_len)
                    if not record_data or len(record_data) < record_len:
                        break
                    
                    if record_data[0] != 0x20:  # Skip deleted records
                        continue
                    
                    record = {}
                    offset = 1  # Skip deletion flag
                    
                    for field in self.fields:
                        field_data = record_data[offset:offset+field['length']]
                        value = field_data.decode('ascii', errors='ignore').strip()
                        record[field['name']] = value
                        offset += field['length']
                    
                    self.records.append(record)
                    
        except Exception as e:
            print(f"Error reading {self.filename}: {str(e)}")

class ExcelReader:
    def __init__(self, filename):
        self.filename = filename
        self.records = []
        self.fields = []
        self.read_excel()
    
    def read_excel(self):
        if not EXCEL_SUPPORT:
            return
            
        try:
            wb = openpyxl.load_workbook(self.filename, read_only=True, data_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
                if not headers:
                    continue
                
                # Store fields info (similar to DBF format)
                if not self.fields:  # Only set fields from first sheet
                    self.fields = [{'name': h, 'type': 'C', 'length': 255} for h in headers]
                
                # Read data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    record = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers):
                            # Convert value to string and handle None
                            str_value = str(value) if value is not None else ''
                            record[headers[col_idx]] = str_value.strip()
                    
                    # Add sheet info to record
                    record['_sheet'] = sheet_name
                    self.records.append(record)
                    
                    # Limit records for performance
                    if len(self.records) >= 50000:
                        break
                
                if len(self.records) >= 50000:
                    break
                    
            wb.close()
            
        except Exception as e:
            print(f"Error reading Excel file {self.filename}: {str(e)}")

class PartLookupApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Offline Part Number Lookup (with Excel Support)")
        self.root.geometry("900x700")
        
        # Data storage
        self.all_data = {}
        self.data_loaded = False
        
        # Get the directory where the executable/script is located
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            self.base_path = os.path.dirname(sys.executable)
        else:
            # Running as script
            self.base_path = os.path.dirname(os.path.abspath(__file__))
        
        # Setup GUI
        self.setup_gui()
        
        # Load data
        self.load_data()
    
    def setup_gui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Search frame
        search_frame = ttk.LabelFrame(main_frame, text="Part Number Search", padding="10")
        search_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Part number entry
        ttk.Label(search_frame, text="Enter Part Number:").grid(row=0, column=0, padx=(0, 10))
        
        self.part_entry = ttk.Entry(search_frame, width=30)
        self.part_entry.grid(row=0, column=1, padx=(0, 10))
        self.part_entry.bind('<Return>', lambda e: self.search_part())
        
        # Search button
        self.search_btn = ttk.Button(search_frame, text="Search", command=self.search_part)
        self.search_btn.grid(row=0, column=2, padx=(0, 10))
        
        # Clear button
        ttk.Button(search_frame, text="Clear", command=self.clear_all).grid(row=0, column=3)
        
        # Status label
        self.status_label = ttk.Label(main_frame, text="Loading data...", foreground="blue")
        self.status_label.grid(row=1, column=0, columnspan=2, pady=5)
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Search Results", padding="10")
        results_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        
        # Results text area
        self.results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, width=80, height=25)
        self.results_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
    
    def load_data(self):
        # Try to find the data directory
        possible_paths = [
            os.path.join(self.base_path, 'AirDataDatabase'),
            os.path.join(self.base_path, 'assets', 'AirDataDatabase'),
            os.path.join(os.path.dirname(self.base_path), 'assets', 'AirDataDatabase'),
            '/var/www/cal.lueshub.com/assets/AirDataDatabase'
        ]
        
        data_dir = None
        for path in possible_paths:
            if os.path.exists(path):
                data_dir = path
                break
        
        if not data_dir:
            self.status_label.config(text="Error: AirDataDatabase folder not found!", foreground="red")
            messagebox.showerror("Error", "Could not find AirDataDatabase folder. Please ensure it's in the same directory as this application.")
            return
        
        self.status_label.config(text=f"Loading data from {data_dir}...")
        self.root.update()
        
        # Priority files for part data
        dbf_files = ['INVENT.DBF', 'POITEM.DBF', 'BUYQUOTE.DBF', 'ALTPART.DBF', 'KIT.DBF']
        excel_files = ['INVENTORIO ACTUAL GENTHRUST.xlsx']
        
        loaded_count = 0
        
        # Load DBF files
        for filename in dbf_files:
            filepath = os.path.join(data_dir, filename)
            if os.path.exists(filepath):
                self.status_label.config(text=f"Loading {filename}...")
                self.root.update()
                
                reader = DBFReader(filepath)
                if reader.records:
                    self.all_data[filename] = {
                        'fields': reader.fields,
                        'records': reader.records,
                        'type': 'DBF'
                    }
                    loaded_count += 1
        
        # Load Excel files
        if EXCEL_SUPPORT:
            for filename in excel_files:
                filepath = os.path.join(data_dir, filename)
                if os.path.exists(filepath):
                    self.status_label.config(text=f"Loading {filename}...")
                    self.root.update()
                    
                    reader = ExcelReader(filepath)
                    if reader.records:
                        self.all_data[filename] = {
                            'fields': reader.fields,
                            'records': reader.records,
                            'type': 'Excel'
                        }
                        loaded_count += 1
        else:
            # Check if Excel files exist but can't be read
            for filename in excel_files:
                filepath = os.path.join(data_dir, filename)
                if os.path.exists(filepath):
                    print(f"Found {filename} but openpyxl is not installed. Skipping Excel file.")
        
        if loaded_count > 0:
            self.data_loaded = True
            excel_msg = " (Excel support enabled)" if EXCEL_SUPPORT else " (Excel support disabled - install openpyxl)"
            self.status_label.config(text=f"Data loaded from {loaded_count} files{excel_msg}. Ready to search.", foreground="green")
            self.part_entry.focus()
        else:
            self.status_label.config(text="Error: Could not load any data files!", foreground="red")
    
    def search_part(self):
        if not self.data_loaded:
            messagebox.showwarning("Warning", "Data is not loaded yet. Please wait.")
            return
        
        part_number = self.part_entry.get().strip().upper()
        if not part_number:
            messagebox.showwarning("Warning", "Please enter a part number.")
            return
        
        self.results_text.delete(1.0, tk.END)
        self.status_label.config(text=f"Searching for part number: {part_number}...", foreground="blue")
        self.root.update()
        
        results = []
        total_matches = 0
        
        # Search through each loaded file
        for filename, data in self.all_data.items():
            file_matches = []
            
            # Find part number fields
            part_fields = []
            for field in data['fields']:
                field_name = field['name'].upper()
                if 'PART' in field_name or 'ITEM' in field_name or 'NUMBER' in field_name or 'PN' in field_name or 'CODIGO' in field_name:
                    part_fields.append(field['name'])
            
            # If no specific part fields found, search all fields
            if not part_fields:
                part_fields = [f['name'] for f in data['fields']]
            
            # Search records
            for record in data['records']:
                match_found = False
                for field in part_fields:
                    if field in record:
                        value = str(record[field]).upper()
                        if part_number in value or value == part_number:
                            match_found = True
                            break
                
                if match_found:
                    file_matches.append(record)
                    total_matches += 1
            
            if file_matches:
                results.append({
                    'filename': filename,
                    'matches': file_matches,
                    'fields': data['fields'],
                    'type': data.get('type', 'Unknown')
                })
        
        # Display results
        if results:
            self.results_text.insert(tk.END, f"Found {total_matches} matches for part number: {part_number}\n")
            self.results_text.insert(tk.END, "=" * 80 + "\n\n")
            
            for result in results:
                self.results_text.insert(tk.END, f"FILE: {result['filename']} ({result['type']})\n")
                self.results_text.insert(tk.END, "-" * 40 + "\n")
                
                for i, record in enumerate(result['matches'], 1):
                    self.results_text.insert(tk.END, f"\nRecord {i}:")
                    if '_sheet' in record:
                        self.results_text.insert(tk.END, f" [Sheet: {record['_sheet']}]")
                    self.results_text.insert(tk.END, "\n")
                    
                    # Display all non-empty fields
                    for field in result['fields']:
                        field_name = field['name']
                        if field_name in record and record[field_name].strip() and field_name != '_sheet':
                            self.results_text.insert(tk.END, f"  {field_name}: {record[field_name]}\n")
                    
                    self.results_text.insert(tk.END, "\n")
                
                self.results_text.insert(tk.END, "\n" + "=" * 80 + "\n\n")
            
            self.status_label.config(text=f"Search complete. Found {total_matches} matches.", foreground="green")
        else:
            self.results_text.insert(tk.END, f"No matches found for part number: {part_number}\n\n")
            self.results_text.insert(tk.END, "Try searching with a partial part number or check the spelling.")
            self.status_label.config(text="No matches found.", foreground="orange")
    
    def clear_all(self):
        self.part_entry.delete(0, tk.END)
        self.results_text.delete(1.0, tk.END)
        self.status_label.config(text="Ready to search.", foreground="green")
        self.part_entry.focus()

def main():
    root = tk.Tk()
    app = PartLookupApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()